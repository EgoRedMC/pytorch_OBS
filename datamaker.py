import numpy as np
import openpyxl as xl
import random as r
import math as m

np.set_printoptions(threshold=66000)


def fill_arr_2dim(arr, pattern):
    n = arr.shape[0]
    if pattern == 0:
        for i in range(n):
            for j in range(n):
                if j < n / 2:
                    arr[i][j] = 1
    if pattern == 1:
        for i in range(n):
            for j in range(n):
                if i < j:
                    arr[i][j] = 1
        for i in range(n - 1):
            add_square(arr, i, i, r.randint(0, 3), 0)
    if pattern == 2:
        fill_arr_2dim(arr, 0)
        for j in range(m.floor((n / 2) / 4)):
            for i in range(n):
                if i >= (j + 1) * 3 and i <= (n - 1) - (j + 1) * 3:
                    arr[i][m.floor((n / 2)) + j] = 1


def add_square(arr, y, x, size, val):
    for i in range(size):
        for j in range(size):
            if x + i >= arr.shape[0] or y + j >= arr.shape[0]:
                continue
            arr[x + i][y + j] = val


def get_data_2dim(arr, img, length):
    for i in range(length):
        x, y = r.randint(0, img.shape[0] - 1), r.randint(0, img.shape[0] - 1)
        arr[i][0] = x
        arr[i][1] = y
        arr[i][2] = img[x][y]


dim = 2  # не изменяется, сделано для потенциального расширения
size = 32
length = 200
pattern = 2  # менять рисунок здесь
real_length = (int)(length * 3 / 2)
img = np.zeros((size, size), dtype=int)  # size должно повторяться dim раз !!
arr = np.zeros((real_length, dim + 1), dtype=int)
#
fill_arr_2dim(img, pattern)  # функция заполнения должна быть с числом, равным dim !!
# while (True):
#     get_data_2dim(arr, img, real_length)
#     if np.mean(arr[:, dim]) > 0.4 and np.mean(arr[:, dim]) < 0.7:
#         break

get_data_2dim(arr, img, real_length)
#
wb = xl.load_workbook("data.xlsx")
sheet = wb['obs']
#
sheet.cell(row=1, column=1).value = "dim"
sheet.cell(row=2, column=1).value = dim
sheet.cell(row=1, column=2).value = "pattern"
sheet.cell(row=2, column=2).value = pattern
sheet.cell(row=1, column=3).value = "length"
sheet.cell(row=2, column=3).value = length
sheet.cell(row=1, column=4).value = "size"
sheet.cell(row=2, column=4).value = size

for i in range(dim):
    sheet.cell(row=3, column=1 + i).value = "x{}".format(i + 1)
sheet.cell(row=3, column=1 + dim).value = "y"
for i in range(real_length):
    for j in range(dim + 1):
        sheet.cell(row=4 + i, column=1 + j).value = arr[i][j]
wb.save("data.xlsx")

#
sheet = wb['obs_valid']
for i in range(size):
    for j in range(size):
        sheet.cell(row=1 + i * size + j, column=1).value = i
        sheet.cell(row=1 + i * size + j, column=2).value = j
        sheet.cell(row=1 + i * size + j, column=3).value = img[i][j]
wb.save("data.xlsx")

if size <= 32:
    print(img)
print(np.mean(arr[:, dim]))
