import torch as t
import torch.nn.utils.prune as pr
import openpyxl as xl
import math as m
import math
import NNet

device = t.device("cuda" if t.cuda.is_available() else "cpu")
#
wb = xl.load_workbook(filename='data.xlsx', read_only=True)
sheet = wb['obs']
#
dim = int(sheet.cell(column=1, row=2).value)
length = int(sheet.cell(column=3, row=2).value)
size = int(sheet.cell(column=4, row=2).value)
input_data = t.zeros(length, 2).to(device)
output_data = t.zeros(length, 1).to(device)
input_validation = t.zeros((int)(length / 2), 2).to(device)
output_validation = t.zeros((int)(length / 2), 1).to(device)
input_true = t.zeros((int)(m.pow(size, dim)), 2).to(device)
output_true = t.zeros((int)(m.pow(size, dim)), 1).to(device)
#
for i in range(length):
    input_data[i][0] = int(sheet.cell(column=1, row=4 + i).value)
    input_data[i][1] = int(sheet.cell(column=2, row=4 + i).value)
    output_data[i][0] = int(sheet.cell(column=3, row=4 + i).value)
for i in range((int)(length / 2)):
    input_validation[i][0] = int(sheet.cell(column=1, row=4 + length + i).value)
    input_validation[i][1] = int(sheet.cell(column=2, row=4 + length + i).value)
    output_validation[i][0] = int(sheet.cell(column=3, row=4 + length + i).value)

## считывание значений для валидации (тратит много времени, потому используется редко)
# sheet = wb['obs_valid']
# for i in range(size):
#     for j in range(size):  # число вложенных циклов долно быть равно dim !!
#         input_true[i * size + j][0] = \
#             int(sheet.cell(column=1, row=1 + i * size + j).value)
#         input_true[i * size + j][1] = \
#             int(sheet.cell(column=2, row=1 + i * size + j).value)
#         output_true[i * size + j][0] = \
#             int(sheet.cell(column=3, row=1 + i * size + j).value)

wb.close()

#
#
net1 = NNet.Net(dim, device, 3, 15, "sigm")
net2 = NNet.copy_net(net1)
epoch_amount = 10000
optimizer1 = t.optim.SGD(net1.parameters(), lr=0.05)
optimizer2 = t.optim.SGD(net2.parameters(), lr=0.05)
loss = t.nn.MSELoss()


def learning(net, optimizer):
    for epoch_number in range(epoch_amount):
        optimizer.zero_grad()
        #
        prediction = net.forward(input_data)
        loss_val = loss(prediction, output_data)
        loss_val.backward()
        optimizer.step()
        # if epoch_number % 500 == 0:
        #     print(loss_val.item())
        #     print("cross valid: "
        #           + str(loss(net.forward(input_validation),
        #                      output_validation).item()))


def net_print(net):
    # net.print_info()
    print("loss value: " + str(loss(net.forward(input_data),
                                    output_data).item()))
    print("cross valid: " + str(loss(net.forward(input_validation),
                                     output_validation).item()))
    # print("cross valid round: " +
    #       str(loss(net.forward_round(input_validation),
    #                output_validation).item()))
    # получение R(alpha_l)
    if not input_true.to("cpu").sum().eq(t.zeros(1)):
        print("true valid: " + str(loss(net.forward(input_true),
                                        output_true).item()))
        print("true valid round: " +
              str(loss(net.forward_round(input_true),
                       output_true).item()))


def eval_hessian(loss_grad, model):
    cnt = 0
    for g in loss_grad:
        g_vector = g.contiguous().view(-1) if cnt == 0 else t.cat([g_vector, g.contiguous().view(-1)])
        cnt = 1
    l = g_vector.size(0)
    hessian = t.zeros(l, l)
    for idx in range(l):
        grad2rd = t.autograd.grad(g_vector[idx], model.parameters(), create_graph=True)
        cnt = 0
        for g in grad2rd:
            g2 = g.contiguous().view(-1) if cnt == 0 else t.cat([g2, g.contiguous().view(-1)])
            cnt = 1
        hessian[idx] = g2
    return hessian.to(device)


def get_indices(arr, lv_item, val):
    pass
    indices = []
    i = -1
    for l in arr:
        i += 1
        if arr[i].item() * val < lv_item:
            # if arr[i].item()  < 0:
            # 10^3 для net.l=1, 10^6 (?,!!!) для net.l=2
            indices.append(i)
    return indices


def get_masks(indices, net):
    layers = net.l
    k = net.k
    masks = [t.ones(k, 2), t.ones(k)]
    cnt = 0
    for i in indices:
        if i - 2 * k < 0:
            masks[0][i // 2][i % 2] = 0
            continue
        i -= 2 * k
        if i - k < 0:
            masks[1][i] = 0
            continue
        i -= k
        if layers >= 2:
            if cnt == 0:
                masks.append(t.ones(k, k))
                masks.append(t.ones(k))
                cnt += 1
            if i - k * k < 0:
                masks[2][i // k][i % k] = 0
                continue
            i -= k * k
            if i - k < 0:
                masks[3][i] = 0
                continue
            i -= k
            if layers >= 3:
                if cnt == 1:
                    masks.append(t.ones(k, k))
                    masks.append(t.ones(k))
                    cnt += 1
                if i - k * k < 0:
                    masks[4][i // k][i % k] = 0
                    continue
                i -= k * k
                if i - k < 0:
                    masks[5][i] = 0
                    continue
                i -= k
        if len(masks) < 2 * layers + 1:
            masks.append(t.ones(1, k))
            masks.append(t.ones(1))
        if i - k < 0:
            masks[-2][0][i] = 0
            continue
        i -= k
        if i - 1 < 0:
            masks[-1][0] = 0
            continue
    if len(masks) < 2 * layers + 1:
        masks.append(t.ones(1, k))
        masks.append(t.ones(1))
    return masks


def fuse_masks(masks1, masks2, masks3=[], masks4=[]):
    res_masks = []
    res_masks.append(masks1[0])
    res_masks.append(masks1[1])
    res_masks.append(masks2[2])
    res_masks.append(masks2[3])
    if not masks3 == []:
        res_masks.append(masks3[4])
        res_masks.append(masks3[5])
        if not masks4 == []:
            res_masks.append(masks4[6])
            res_masks.append(masks4[7])
    return res_masks


def prune_by_masks(net, masks):
    pr.identity(net.l1, "weight")
    pr.identity(net.l1, "bias")
    net.l1.weight_mask = masks[0].to(device)
    net.l1.bias_mask = masks[1].to(device)
    if net.l >= 2:
        pr.identity(net.l2, "weight")
        pr.identity(net.l2, "bias")
        net.l2.weight_mask = masks[2].to(device)
        net.l2.bias_mask = masks[3].to(device)
        if net.l >= 3:
            pr.identity(net.l3, "weight")
            pr.identity(net.l3, "bias")
            net.l3.weight_mask = masks[4].to(device)
            net.l3.bias_mask = masks[5].to(device)
    pr.identity(net.out, "weight")
    pr.identity(net.out, "bias")
    net.out.weight_mask = masks[-2].to(device)
    net.out.bias_mask = masks[-1].to(device)


def make_masks(arr, net, lv_item, val1, val2, val3=0, val4=0):
    prune_ind1 = get_indices(arr, lv_item, val1)
    masks1 = get_masks(prune_ind1, net)
    prune_ind2 = get_indices(arr, lv_item, val2)
    masks2 = get_masks(prune_ind2, net)
    if not val3 == 0:
        prune_ind3 = get_indices(arr, lv_item, val3)
        masks3 = get_masks(prune_ind3, net)
        if not val4 == 0:
            prune_ind4 = get_indices(arr, lv_item, val4)
            masks4 = get_masks(prune_ind4, net)
            return fuse_masks(masks1, masks2, masks3, masks4)
        return fuse_masks(masks1, masks2, masks3)
    return fuse_masks(masks1, masks2)


def l_print(L_arr, n_count, l):
    arr = []
    cur_l = 1
    for i in L_arr:
        arr.append(round(math.log10(i.item())))
        if cur_l == 1:
            if len(arr) == n_count * 2 + n_count:
                for j in range(0, 3):
                    print(arr[j * n_count:(j + 1) * n_count])
                print()
                arr = []
                cur_l += 1
        if cur_l > 1 and cur_l < l + 1:
            if len(arr) == n_count * n_count + n_count:
                for j in range(0, n_count + 1):
                    print(arr[j * n_count:(j + 1) * n_count])
                print()
                arr = []
                cur_l += 1
        if cur_l == l + 1:
            if len(arr) == n_count + 1:
                print(arr[0:n_count])
                print(arr[n_count])
                print()
                arr = []
                cur_l += 1


learning(net1, optimizer1)
net_print(net1)
print()
#
optimizer1.zero_grad()
prediction = net1.forward(input_data)
loss_val = loss(prediction, output_data)
grads = t.autograd.grad(loss_val, net1.parameters(), create_graph=True)
hess = eval_hessian(grads, net1)
hess_inv = t.inverse(
    hess)  # расчет обратного гессиана осуществляется напрямую, без использоования оптимального алгоритма (потери во времени и памяти)
diag = t.diagonal(hess_inv, 0)
cnt = 0
for g in net1.parameters():
    net_squeezed = g.contiguous().view(-1) if cnt == 0 else t.cat([net_squeezed, g.contiguous().view(-1)])
    cnt = 1
L_arr = t.abs(t.div(t.square(net_squeezed), diag))
# print(L_arr)
# masks = make_masks(L_arr, net1, loss_val.item(),  # для OBS - значения для всех слоев одинаковы
#                    5 * (10 ** 10),
#                    5 * (10 ** 10),
#                    5 * (10 ** 10),
#                    5 * (10 ** 10))
masks = make_masks(L_arr, net1, loss_val.item(),  # для L-OBS - значения для каждого слоя своё
                   5 * (10 ** 3),
                   5 * (10 ** 10),
                   1 * (10 ** 10),
                   5 * (10 ** 2))
for m in masks:
    print("Число единиц в маске: " + str(m.sum()) + " ; Форма маски: " + str(m.shape))
    pass
print()
# l_print(L_arr, net1.k, net1.l)
prune_by_masks(net1, masks)
net_print(net1)
